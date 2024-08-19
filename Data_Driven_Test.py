
"""
pip install openpyxl
#import pytest

"""
import time
import openpyxl
import unittest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.support.ui import Select
book = openpyxl.load_workbook("C:\\Users\\Ahmed Shahin\\OneDrive\\Desktop\\my folder\\Test1.xlsx")
sheet = book['Sheet1']
driver = webdriver.Chrome()
class Test(unittest.TestCase):
 def perform_login(username):
  block1 = sheet.cell(row = 1, column = 2)
  block2 = sheet.cell(row = 2, column = 2)
  block3 = sheet.cell(row = 3, column = 2)
  block4 = sheet.cell(row = 4, column = 2)
  block5 = sheet.cell(row = 5, column = 2)
  block6 = sheet.cell(row = 6, column = 2)
  block7 = sheet.cell(row = 7, column = 2)
  block8 = sheet.cell(row = 8, column = 2)
  block9 = sheet.cell(row = 9, column = 2)
  block10 = sheet.cell(row = 10, column = 2)
  block11 = sheet.cell(row = 11, column = 2)
  block12 = sheet.cell(row = 12, column = 2)
  block13 = sheet.cell(row = 13, column = 2)
  block14 = sheet.cell(row = 14, column = 2)
  block15 = sheet.cell(row = 15, column = 2)
  block16 = sheet.cell(row = 16, column = 2)
  block17 = sheet.cell(row = 17, column = 2)
  url1 = "https://app.jackrabbitclass.com/regv2.asp?id=530822"
  url2 = "https://www.google.com/"
  url3 = "https://app.jackrabbitclass.com/jr3.0/ParentPortal/Login?orgId=530822"
  # driver.get(url2)
  # driver.maximize_window()
  # driver.execute_script("window.open('about:blank','secondtab');")
  # driver.switch_to.window("secondtab")
  driver.get(url1)
  driver.execute_script("window.scrollTo(0,400);")
  ddelement= Select(driver.find_element(By.XPATH, "/html/body/div[3]/form/fieldset[1]/div/div[1]/select"))
  ddelement.select_by_index(2)
  time.sleep(0.3)
  driver.find_element(By.ID, "ReferralName").send_keys(block1.value) 
  time.sleep(0.3)
  driver.find_element(By.XPATH, "/html/body/div[3]/form/fieldset[2]/div[1]/div/input").send_keys(block2.value)
  time.sleep(0.3)
  driver.find_element(By.NAME, "Addr").send_keys(block3.value) 
  time.sleep(0.3)
  driver.find_element(By.NAME, "City").send_keys(block4.value) 
  time.sleep(0.3)
  element= Select(driver.find_element(By.ID, "State"))
  element.select_by_index(2)
  time.sleep(0.3)
  driver.find_element(By.NAME, "Zip").send_keys(block5.value) 
  time.sleep(0.3)
  driver.find_element(By.NAME, "HPhone").send_keys(block6.value) 
  time.sleep(0.3)
  driver.execute_script("window.scrollTo(0,1100);")
  time.sleep(0.3)
  driver.find_element(By.NAME, "MFName").send_keys(block7.value) 
  time.sleep(0.3)
  driver.find_element(By.NAME, "MLName").send_keys(block8.value) 
  time.sleep(0.3)
  relation= Select(driver.find_element(By.XPATH, "/html/body/div[3]/form/fieldset[3]/div[1]/div[3]/select"))
  relation.select_by_index(2)
  time.sleep(0.3)
  driver.find_element(By.NAME, "MHPhone").send_keys(block9.value)
  time.sleep(0.3)
  driver.find_element(By.NAME, "MEmail").send_keys(block10.value)
  time.sleep(0.3)
  driver.find_element(By.NAME, "ConfirmMEmail").send_keys(block11.value)
  time.sleep(0.3)
  driver.find_element(By.NAME, "MEPassword").send_keys(block12.value)
  time.sleep(0.3)
  driver.find_element(By.NAME, "MEConfirmPassword").send_keys(block13.value)
  time.sleep(0.3)
  driver.execute_script("window.scrollTo(0,1650);")
  driver.find_element(By.NAME, "S1FName").send_keys(block14.value)
  time.sleep(0.3)
  driver.find_element(By.NAME, "S1LName").send_keys(block15.value)
  time.sleep(0.3)
  gender= Select(driver.find_element(By.NAME, "S1Gender"))
  gender.select_by_index(2)
  time.sleep(0.3)
  driver.find_element(By.NAME, "S1BDate").send_keys(block16.value)
  time.sleep(0.3)
  ShirtSize= Select(driver.find_element(By.NAME, "S1TShirtSize"))
  ShirtSize.select_by_index(2)
  time.sleep(0.3)
  gradelevel= Select(driver.find_element(By.NAME, "S1Grade"))
  gradelevel.select_by_index(2)
  time.sleep(0.3)
  driver.find_element(By.NAME, "S1Allergies").send_keys(block17.value)
  time.sleep(0.3)
  # driver.execute_script("window.open('about:blank','thirdtab');")
  # driver.switch_to.window("thirdtab")
  # driver.get(url1)
  # actions = ActionChains(driver)
  # actions.perform()
  # name = driver.find_element(By.XPATH, "/html/body/div/div[2]/div/div/input")
  # actions.move_to_element(name).perform()
  # name.click()
  print(" Test passed  ")
  driver.get(url3)
  time.sleep(2)
  driver.quit()       
Data= Test()
Data.perform_login()
















"""
class Test(unittest.TestCase):

    def test_1(shop):
       
        url = "https://www.google.com"
        driver.get(url)
        driver.maximize_window()
        actions = ActionChains(driver)
        driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/form/div[1]/div[1]/div[1]/div/div[2]/textarea").send_keys("Etisalat")
        #actions.send_keys(keys.ENTER)
        actions.perform()
        driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/form/div[1]/div[1]/div[4]/center/input[1]").click()
        driver.find_element(By.XPATH, "//*[contains(text(),'e& (etisalat and ) UAE | Mobile plans, TV & Internet plans, WiFi ...')]").click()
        driver.find_element(By.XPATH, "//*[contains(text(),'Devices')]").click()
        Samsung=driver.find_element(By.XPATH, "/html/body/div[1]/div/div/div[2]/div/div/div/div[2]/div[1]/nav/div/div/div[3]/div/div[1]/div/div/ul/li[3]/div/section/div/div/div[2]/div[2]/div[2]/a")
        actions.move_to_element(Samsung).perform()
        Samsung.click()
        driver.execute_script("window.scrollTo(0,300);")
        phones=driver.find_element(By.XPATH, "/html/body/div[1]/div/div/div[3]/div/div/div[3]/div/section/div[1]/div/div[1]/div/ul/li[2]")
        actions.move_to_element(phones).perform()
        phones.click()
        driver.execute_script("window.scrollTo(0,700);")
        time.sleep(2)
        S24=driver.find_element(By.XPATH, "/html/body/div[1]/div/div/div[3]/div/div/div[3]/div/div[2]/section/div/div/div[2]/div/div/div[3]/div/div/a")
        actions.move_to_element(S24).perform()
        S24.click()
        time.sleep(2)
        storage=driver.find_element(By.XPATH, "//*[contains(text(),'256GB')]")
        actions.move_to_element(storage).perform()
        storage.click()
        time.sleep(2)
        driver.execute_script("window.scrollTo(0,700);")
        time.sleep(2)
        driver.find_element(By.ID, "configure").click
        #CART=driver.find_element(By.XPATH, "/html/body/div[1]/ui-view/device-configuration-component/section/div/div/div[3]/price-details/div/div/div[3]/div/div[2]/ng-container[7]/div/div[2]")
        #driver.execute_script("window.scrollTo(0,300);")
        #actions.move_to_element(CART).perform()
        time.sleep(2)
        #CART.click()
        time.sleep(2)
        last=driver.find_element(By.XPATH, "//*[contains(text(),'S24')]")
        print(last.text," Test passed  ")
        shop.assertEqual(last.text, "")
        time.sleep(2)
        driver.quit() 
    
shop = Test()
shop.test_1()

"""

# print(last.text," please run the (python -m unittest Etisalat_shop.py) command to run the TC to run it comment the last 2 lines  ")